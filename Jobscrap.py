import sys
import pandas as pd
from selenium import webdriver
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.wait import WebDriverWait
from openpyxl import load_workbook
pd.option_context('display.max_colwidth', 0)

options = webdriver.ChromeOptions()
# options.add_argument("user-agent=Chrome/87.0.4280.88")
options.add_argument('headless')
options.add_argument('--window-size=1920,1080')
driver = webdriver.Chrome(executable_path="D:\Drivers\chromedriver.exe", options=options)
# driver.delete_all_cookies()
# check for keywords and pass links
for i in range(1,len(sys.argv)):
    if sys.argv[i] == 'Selenium':
        url = 'https://www.glassdoor.com/Job/selenium-jobs-SRCH_KO0,8.htm?&jobType=fulltime&fromAge=1&includeNoSalaryJobs=true'
    elif sys.argv[i] == 'API Testing':
        url='https://www.glassdoor.com/Job/api-testing-jobs-SRCH_KO0,11.htm?jobType=fulltime&fromAge=1&includeNoSalaryJobs=true'
    elif sys.argv[i] == 'QA Automation':
        url='https://www.glassdoor.com/Job/qa-automation-jobs-SRCH_KO0,13.htm?jobType=fulltime&fromAge=1&includeNoSalaryJobs=true'
    driver.get(url)
    start_time = time.time()
    # driver.maximize_window()

    time.sleep(1)
    # check if results are loaded
    myElement = driver.find_elements_by_xpath('//*[@id="MainCol"]/div[1]/div[2]/div/div[1]/p')
    if len(myElement) != 0:
        print(myElement[0].text)
    else:
        # extract number of pages
        m = driver.find_element_by_xpath("//*[@id='MainCol']/div[2]/div[1]").text
        numPages = int(m.split(' ')[3])
        # check for popup
        isAvailable = len(driver.find_elements_by_xpath("//*[@id='JAModal']/div/div[2]"))
        if isAvailable > 0:
            print('yes')
            if driver.find_element_by_xpath("//*[@id='JAModal']/div/div[2]").is_displayed():
                driver.find_element_by_xpath("//*[@id='JAModal']/div/div[2]/span").click()
        else:
            print('no')
        # initialise first page
        iPage: int = 1
        # jobs empty list
        jobs = []
        # Check if value of iPage is less than numPages and fetch jobs
        while iPage <= numPages:
            if iPage > 1:
                break
            print(driver.find_element_by_xpath("//*[@id='MainCol']/div[2]/div[1]").text)
            # Fetch all jobs from left side into webelement list
            job_buttons = driver.find_elements_by_xpath("//*[@id='MainCol']/div[1]/ul/li")
            num_jobs = len(job_buttons)
            # if first page then total jobs is recent records else count of entire list
            if iPage == 1:
                total_jobs = len(jobs)
            else:
                total_jobs = total_jobs + len(jobs)

            if job_buttons != 0:
                msg = "{0} jobs available."
                print(msg.format(num_jobs))
                # Wait till page is loaded
                myElem = WebDriverWait(driver, 50).until(ec.presence_of_element_located((By.XPATH, '//*[@id="MainCol"]')))
                print('Page is ready')
                # driver.find_element_by_xpath('//*[@id="MainCol"]').click()
                # Loop thru every job link and fetch details
                for j in range(num_jobs):
                    print("Progress: {}".format("" + str(j) + "/" + str(num_jobs)))

                   # jobs.append('Hi')
                   #Check for popup
                    isAvailable = len(driver.find_elements_by_xpath("//*[@id='JAModal']/div/div[2]"))
                    if isAvailable > 0:
                       print('yes popup')
                       if driver.find_element_by_xpath("//*[@id='JAModal']/div/div[2]").is_displayed():
                          driver.find_element_by_xpath("//*[@id='JAModal']/div/div[2]/span").click()
                    else:
                        print('no popup')
                    time.sleep(2)
                    job_link = job_buttons[j].find_element_by_xpath('.//div/a').get_attribute(
                          'href')
                    print('Job Link: ' + job_link)
                    jobId = job_buttons[j].get_attribute('data-id')
                    print('JobId: ' + jobId)
                    job_buttons[j].click()
                    # Wait for job desc to load
                    WebDriverWait(driver, 50).until(ec.presence_of_element_located((By.CLASS_NAME, 'css-vwxtm')))
                    JobInfo = driver.find_element_by_class_name("css-vwxtm")

                    section = JobInfo.find_element_by_class_name("css-19txzrf")

                    # print(section.get_attribute("innerHTML"))
                    company = section.find_element_by_class_name("css-87uc0g").get_attribute("innerText")
                    # print('Company: ' + company)
                    title = section.find_element_by_class_name("css-1vg6q84").get_attribute("innerText")
                    # print('Job Title: ' + title)
                    location = section.find_element_by_class_name("css-56kyx5").get_attribute("innerText")
                    # print(location)

                    JobDesc = driver.find_element_by_class_name("css-1bqnmih")
                    jobDetails = JobDesc.find_element_by_class_name("jobDescriptionContent").get_attribute("innerText")


                    # Append jobs details into jobs list
                    jobs.append({"Job Id": jobId, "Job Title": title, "Company Name": company,
                                 "Location": location, "Job Description": jobDetails,"Job Link": job_link})
                    # Check for popups
                    isAvailable = len(driver.find_elements_by_xpath("//*[@id='JAModal']/div/div[2]"))
                    if isAvailable > 0:
                        print('yes popup')
                        if driver.find_element_by_xpath("//*[@id='JAModal']/div/div[2]").is_displayed():
                            driver.find_element_by_xpath("//*[@id='JAModal']/div/div[2]/span").click()
                    else:
                        print('no popup')
                    # Check for next button in page
                nCnt: int = len(driver.find_elements_by_xpath("//*[@id='FooterPageNav']/div/ul/li[contains(@class, 'css-114lpwu')]"))
                # if next button is available then click
                if nCnt > 0:
                    bNext = True
                    print('Next is available')
                    driver.find_element_by_xpath("//*[@id='FooterPageNav']/div/ul/li[contains(@class, 'css-114lpwu')]").click()
                    print('Clicking Next Button')
                    time.sleep(2)
                    isAvailable = len(driver.find_elements_by_xpath("//*[@id='JAModal']/div/div[2]"))
                    if isAvailable > 0:
                       print('yes popup')
                       if driver.find_element_by_xpath("//*[@id='JAModal']/div/div[2]").is_displayed():
                            driver.find_element_by_xpath("//*[@id='JAModal']/div/div[2]/span").click()
                    else:
                       print('no popup')
                    # come out of while loop
                else:
                    bNext = False
                    break
            # Increment page till total pages reached
            iPage = iPage + 1

        total_jobs = len(jobs)
        print('Total jobs scrapped: ' + str(total_jobs))
        print('End of Jobs')

        print("--- %s seconds ---" % (time.time() - start_time))
        # Convert list to dataframe
        df = pd.DataFrame(jobs)
        # Trim spaces in all columns
        df['Job Id'] = df['Job Id'].str.strip()
        df['Job Title'] = df['Job Title'].str.strip()
        df['Company Name'] = df['Company Name'].str.strip()
        df['Location'] = df['Location'].str.strip()
        df['Job Description'] = df['Job Description'].str.strip()
        df['Job Link'] = df['Job Link'].str.strip()

        print(df)
            # df.to_csv('my_csv.csv', mode='a', header=False)
        # Code to write or append into excel
        book = load_workbook('C:/Users/Sukanya Ruthvik/PycharmProjects/JobScrappingProject/test.xlsx')
        writer = pd.ExcelWriter('C:/Users/Sukanya Ruthvik/PycharmProjects/JobScrappingProject/test.xlsx', engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        for sheetname in writer.sheets:

            if writer.sheets[sheetname].max_row == 1:
                df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index=False, header=True)
            else:
                df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)
        writer.save()
driver.close()
driver.quit()
# df.to_excel('D:\df.xlsx')
